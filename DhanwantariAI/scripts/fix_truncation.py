"""
fix_truncation.py
-----------------
Reads the complete text from V1-V10 Excel sheets and patches ALL truncated
text fields in the enriched JSON.

Excel layout (confirmed by inspection):
  - One sheet per disease (150 sheets across V1-V10)
  - Row 1:  column headers
  - Rows 2+: bullet points, one per row, in each column

Column → JSON field mapping:
  'India-Specific Complications & Risk Factors'  → india_specific
  'Specific Nutritional Deficiencies in India'   → india_specific  (malnutrition only)
  'Important Considerations'                     → important_notes
  'Confirmation Tests'                           → tests
  'Generic Medicines'                            → generic_medicines
  'Jan Aushadhi Medicines'                       → janaushadhi_medicines
  'Ayurvedic Medicines'                          → ayurvedic_medicines

Run:
  python3 scripts/fix_truncation.py
Output:
  DhanwantariAI_Symptom_Disease_Mapping_FIXED.json
"""

import json
from pathlib import Path

BASE = Path('/Volumes/SatyBkup/DhanwantariDataset')

try:
    import openpyxl
except ImportError:
    raise SystemExit("Run: pip install openpyxl")

# ── Column header → JSON field key ──────────────────────────────────────────
# Keys are lowercased for case-insensitive matching.
HEADER_TO_FIELD = {
    'india-specific complications & risk factors': 'india_specific',
    'specific nutritional deficiencies in india':  'india_specific',
    'important considerations':                    'important_notes',
    'confirmation tests':                          'tests',
    'generic medicines':                           'generic_medicines',
    'jan aushadhi medicines':                      'janaushadhi_medicines',
    'ayurvedic medicines':                         'ayurvedic_medicines',
}

EXCEL_FILES = [f'DhanwantariAI_Disease_Dataset_V{i}.xlsx' for i in range(1, 11)]

# ── Load enriched JSON ──────────────────────────────────────────────────────
json_path = BASE / 'DhanwantariAI_Symptom_Disease_Mapping (1).json'
print(f'Loading JSON: {json_path.name}')
with open(json_path, encoding='utf-8') as f:
    data = json.load(f)

# Build lookup: disease name (lower) → disease object
disease_map = {d['name'].lower().strip(): d for d in data['diseases']}
print(f'Diseases in JSON: {len(disease_map)}')

# ── Per-field patch counters ─────────────────────────────────────────────────
field_patch_counts = {k: 0 for k in set(HEADER_TO_FIELD.values())}
not_found = []
diseases_touched = set()


def collect_column(rows: list, col_idx: int) -> str:
    """
    Collect all non-empty cell values from rows[1:] (skipping header row)
    in column col_idx, and join them with double newlines.
    Each row = one bullet point in the original sheet.
    """
    items = []
    for row in rows[1:]:
        if col_idx < len(row) and row[col_idx] is not None:
            val = str(row[col_idx]).strip()
            if val:
                items.append(val)
    return '\n\n'.join(items)


# ── Main loop ────────────────────────────────────────────────────────────────
for filename in EXCEL_FILES:
    path = BASE / filename
    if not path.exists():
        print(f'  MISSING: {filename}')
        continue

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # ── Match sheet → JSON disease ───────────────────────────────────
        key = sheet_name.lower().strip()
        if key not in disease_map:
            not_found.append(f'{filename}::{sheet_name}')
            continue

        disease = disease_map[key]

        # ── Build col_idx → json_field map from header row ───────────────
        header_row = rows[0]
        col_field_map: dict[int, str] = {}
        for ci, cell in enumerate(header_row):
            if cell is None:
                continue
            label = str(cell).strip().lower()
            if label in HEADER_TO_FIELD:
                col_field_map[ci] = HEADER_TO_FIELD[label]

        # ── Extract full text per field and patch if longer ───────────────
        for col_idx, json_field in col_field_map.items():
            excel_text = collect_column(rows, col_idx)
            if not excel_text:
                continue

            current = disease.get(json_field, '') or ''
            if len(excel_text) > len(current):
                disease[json_field] = excel_text
                field_patch_counts[json_field] += 1
                diseases_touched.add(key)

    wb.close()

# ── Summary ──────────────────────────────────────────────────────────────────
print(f'\n{"="*55}')
print(f'Diseases patched:  {len(diseases_touched)} / {len(disease_map)}')
print(f'\nPatch counts per field:')
for field, count in sorted(field_patch_counts.items()):
    print(f'  {field:30s}: {count}')

if not_found:
    print(f'\nSheets not matched to any disease ({len(not_found)}):')
    for x in not_found[:20]:
        print(f'  {x}')

# ── Save output ───────────────────────────────────────────────────────────────
out = BASE / 'DhanwantariAI_Symptom_Disease_Mapping_FIXED.json'
with open(out, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print(f'\nSaved: {out.name}  ({out.stat().st_size / 1_048_576:.1f} MB)')

# ── Verification: check for still-truncated fields ───────────────────────────
print(f'\n{"="*55}')
print('Post-patch truncation check (fields not ending in .!?:)")...')
SENTINEL_FIELDS = ('india_specific', 'important_notes', 'tests',
                   'generic_medicines', 'janaushadhi_medicines', 'ayurvedic_medicines')
still_bad: dict[str, list[str]] = {f: [] for f in SENTINEL_FIELDS}
avg_lengths: dict[str, list[int]] = {f: [] for f in SENTINEL_FIELDS}

for d in data['diseases']:
    for field in SENTINEL_FIELDS:
        val = (d.get(field) or '').strip()
        if val:
            avg_lengths[field].append(len(val))
            if val[-1] not in '.!?:"\'':
                still_bad[field].append(d['name'])

print(f'\n{"Field":<35} {"Avg chars":>10}  {"Still truncated":>15}')
print('-' * 65)
for field in SENTINEL_FIELDS:
    lengths = avg_lengths[field]
    avg = int(sum(lengths) / len(lengths)) if lengths else 0
    bad = len(still_bad[field])
    flag = '  ← CHECK' if bad > 5 else ''
    print(f'  {field:<33} {avg:>10,}  {bad:>15}{flag}')

total_bad = sum(len(v) for v in still_bad.values())
print(f'\nTotal still-truncated fields: {total_bad}  (was 281 before fix)')
print('\nDone. Use DhanwantariAI_Symptom_Disease_Mapping_FIXED.json for fine-tuning.')
